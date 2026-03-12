import openpyxl
import json
import shutil
import os

# Copy file first to avoid lock issues (file may be open in Excel)
src = r'C:\Users\iraki\OneDrive\Desktop\pms\PMS\data\Assets reviewed.xlsx'
tmp = r'C:\Users\iraki\OneDrive\Desktop\pms\PMS\scripts\temp_assets.xlsx'
shutil.copy2(src, tmp)
wb = openpyxl.load_workbook(tmp)
ws = wb.active

print(f"Total data rows: {ws.max_row - 1}")

# Build header map
headers = {}
for cell in ws[1]:
    headers[cell.column] = cell.value

# Column indices
col_assetnum = None
col_desc = None
col_location = None
col_assettype = None
col_serial = None
col_modeltype = None

for col_idx, col_name in headers.items():
    if col_name == 'ASSETNUM':
        col_assetnum = col_idx
    elif col_name == 'DESCRIPTION':
        col_desc = col_idx
    elif col_name == 'LOCATION':
        col_location = col_idx
    elif col_name == 'ASSETTYPE':
        col_assettype = col_idx
    elif col_name == 'SERIALNUM':
        col_serial = col_idx
    elif col_name == 'XMODELTYPE':
        col_modeltype = col_idx

out_path = r'C:\Users\iraki\OneDrive\Desktop\pms\PMS\scripts\seed_assets.sql'

def sql_escape(val):
    if val is None:
        return 'NULL'
    s = str(val).replace("'", "''")
    return f"'{s}'"

with open(out_path, 'w', encoding='utf-8') as f:
    f.write("-- Seed data from Assets reviewed.xlsx\n")
    f.write("-- Auto-generated\n\n")
    
    batch_size = 500
    row_count = 0
    batch = []
    
    for row_idx in range(2, ws.max_row + 1):
        kks = ws.cell(row=row_idx, column=col_assetnum).value
        name = ws.cell(row=row_idx, column=col_desc).value
        loc = ws.cell(row=row_idx, column=col_location).value
        atype = ws.cell(row=row_idx, column=col_assettype).value
        serial = ws.cell(row=row_idx, column=col_serial).value
        model = ws.cell(row=row_idx, column=col_modeltype).value
        
        if not name:
            continue
        
        vals = f"  ({sql_escape(kks)}, {sql_escape(name)}, {sql_escape(loc or 'N/A')}, {sql_escape(atype)}, {sql_escape(serial)}, {sql_escape(model)})"
        batch.append(vals)
        row_count += 1
        
        if len(batch) >= batch_size:
            f.write(f"INSERT INTO assets (kks_code, name, location, asset_type, serial_number, model_type) VALUES\n")
            f.write(",\n".join(batch))
            f.write("\nON CONFLICT (kks_code) DO NOTHING;\n\n")
            batch = []
    
    if batch:
        f.write(f"INSERT INTO assets (kks_code, name, location, asset_type, serial_number, model_type) VALUES\n")
        f.write(",\n".join(batch))
        f.write("\nON CONFLICT (kks_code) DO NOTHING;\n\n")
    
    print(f"Written {row_count} rows to {out_path}")

# Cleanup temp file
os.remove(tmp)
print("Done!")
